from __future__ import annotations

import ast
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_SHEET = "BaseData"
SET_SHEET = "วัสดุทั้งหมด"

SIZE_COL = "ขนาดเสา (m)"
HEAD_COL = "รหัสหัวเสา"
MATERIAL_COL = "รายการวัสดุ"
CODE_COL = "รหัสพัสดุ"
QTY_COL = "จำนวน"
TOTAL_COL = "จำนวนรวม"

SET_COL = "Set"
SET_DESC_COL = "คำอธิบาย"
SET_INSTALL_COL = "ติดตั้ง"

WIRE_MATERIALS = {
    "50 PIC": ("PREFORMED DEAD END,FOR AL PARTIALLY INSULATED CONDUCTOR 22 KV. 50 SQ.MM.", "1020250001", "50", True),
    "95 PIC": ("PREFORMED DEAD END,FOR AL PARTIALLY INSULATED CONDUCTOR 22 KV. 95 SQ.MM.", "1020250002", "95", True),
    "185 PIC": ("PREFORMED DEAD END,FOR AL PARTIALLY INSULATED CONDUCTOR 22 KV. 185 SQ.MM.", "1020250004", "185", True),
    "50 SAC": ("PREFORMED D/E,SAC 22kV 50sq.mm. 21.80mm", "1020260202", "50", True),
    "185 SAC": ("PREFORMED D/E,SAC 22kV 185sq.mm. 29.78mm", "1020260205", "185", True),
    "50 A": ("CLAMP,STRAIN,STRAIGHT TYPE,AL 35-70 sq.mm.ACSR 35-50 SQ.MM.", "1030110000", "50", False),
    "50 ACSR": ("CLAMP,STRAIN,STRAIGHT TYPE,AL 35-70 sq.mm.ACSR 35-50 SQ.MM.", "1030110000", "50", False),
    "185 ACSR": ("CLAMP,STRAIN,STRAIGHT TYPE FOR ACSR.120-185 sq.mm.", "1030110007", "185", False),
    "185 A": ("CLAMP,STRAIN,STRAIGHT TYPE,FOR AL 185 SQ.MM.", "1030110004", "185", False),
}

TENSIONLESS_MATERIALS = {
    "50": ("CONNECTOR,SPLICE,COMPRESSION TYPE,TENSIONLESS AL 50 SQ.MM.", "1020410002"),
    "185": ("CONNECTOR,SPLICE,COMPRESSION TYPE,TENSIONLESS AL 185 SQ.MM.", "1020410027"),
}

PG3_MATERIAL = (
    "CONNECTOR,PARALLEL GROOVE,TRIPLE BOLT,AL,AL-ALLOY AND ACSR 70-185 SQ.MM.",
    "1020300103",
)
HOTLINE_CLAMP_MATERIAL = ("HOTLINE CLAMP,MAIN35-185,TAP50-185SQ.MM.", "1020330104")
BAIL_CLAMP_MATERIAL = ("HOTLINE BAIL-CLAMP,MAIN 70-185 SQ.MM.", "1020330006")
CLEVIS_MATERIAL = ("CLEVIS,THIMBLE,FOR PREFORMED DEAD-END", "1030140011")
TENSIONLESS_TAPES = (
    ("PVC TAPE", "1020180001"),
    ("ERP TAPE", "1020180008"),
)


class MaterialWorkbook:
    def __init__(self) -> None:
        self.base_df: pd.DataFrame | None = None
        self.set_df: pd.DataFrame | None = None
        self.summary: list[dict[str, Any]] = []
        self.base_path: Path | None = None
        self.set_path: Path | None = None

    def load_base(self, path: str | Path) -> dict[str, Any]:
        df = pd.read_excel(path, sheet_name=BASE_SHEET)
        require_columns(df, [SIZE_COL, HEAD_COL, MATERIAL_COL, CODE_COL, QTY_COL], "BaseData")
        df = df[[SIZE_COL, HEAD_COL, MATERIAL_COL, CODE_COL, QTY_COL]].copy()
        df = df.dropna(subset=[SIZE_COL, HEAD_COL, CODE_COL])
        self.base_df = df
        self.base_path = Path(path)
        self.summary = []
        return {
            "file": self.base_path.name,
            "rows": int(len(df)),
            "sizes": self.get_sizes(),
        }

    def load_set(self, path: str | Path) -> dict[str, Any]:
        df = read_set_sheet(path)
        require_columns(df, [SET_COL, CODE_COL, SET_DESC_COL, SET_INSTALL_COL], "SET")
        df = df[[SET_COL, CODE_COL, SET_DESC_COL, SET_INSTALL_COL]].copy()
        df = df.dropna(subset=[SET_COL, CODE_COL])
        self.set_df = df
        self.set_path = Path(path)
        return {
            "file": self.set_path.name,
            "rows": int(len(df)),
            "sets": int(df[SET_COL].astype(str).str.strip().str.lower().nunique()),
        }

    def get_sizes(self) -> list[str]:
        if self.base_df is None:
            return []
        sizes = self.base_df[SIZE_COL].dropna().astype(str).str.strip().unique().tolist()
        return sorted(sizes, key=natural_key)

    def get_heads(self, size: str) -> list[str]:
        if self.base_df is None:
            raise ValueError("ยังไม่ได้โหลดไฟล์ BaseData")
        selected = str(size).strip()
        matches = self.base_df[self.base_df[SIZE_COL].astype(str).str.strip() == selected]
        heads = matches[HEAD_COL].dropna().astype(str).str.strip().unique().tolist()
        return sorted(heads, key=natural_key)

    def get_status(self) -> dict[str, Any]:
        base = None
        if self.base_df is not None and self.base_path is not None:
            base = {
                "file": self.base_path.name,
                "rows": int(len(self.base_df)),
                "sizes": self.get_sizes(),
            }

        set_data = None
        if self.set_df is not None and self.set_path is not None:
            set_data = {
                "file": self.set_path.name,
                "rows": int(len(self.set_df)),
                "sets": int(self.set_df[SET_COL].astype(str).str.strip().str.lower().nunique()),
            }

        return {"base": base, "set": set_data}

    def calculate(self, pages: list[list[dict[str, Any]]]) -> dict[str, Any]:
        if self.base_df is None:
            raise ValueError("ยังไม่ได้โหลดไฟล์ BaseData")

        totals: dict[tuple[str, str], dict[str, Any]] = {}
        input_count = 0
        matched_rows = 0

        for page_number, page in enumerate(pages, start=1):
            for row_number, item in enumerate(page, start=1):
                size = str(item.get("size", "")).strip()
                head = str(item.get("head", "")).strip()
                count = parse_number(item.get("count"))
                if not size or not head or count <= 0:
                    continue

                input_count += 1
                wire_kind = classify_wire_head(head)
                wire1 = clean_text(item.get("wire1"))
                wire2 = clean_text(item.get("wire2"))
                validate_wire_selection(wire_kind, wire1, wire2, page_number, row_number, head)
                matches = self.base_df[
                    (self.base_df[SIZE_COL].astype(str).str.strip() == size)
                    & (self.base_df[HEAD_COL].astype(str).str.strip() == head)
                ]

                for _, row in matches.iterrows():
                    material = clean_text(row[MATERIAL_COL])
                    code = clean_text(row[CODE_COL])
                    amount = parse_number(row[QTY_COL]) * count
                    if not code or amount == 0:
                        continue
                    add_material(totals, material, code, amount)
                    matched_rows += 1

                matched_rows += add_wire_materials(totals, wire_kind, wire1, wire2, count)

        self.summary = sorted(totals.values(), key=lambda r: (str(r[CODE_COL]).lower(), str(r[MATERIAL_COL]).lower()))
        return {
            "items": self.summary,
            "inputRows": input_count,
            "matchedRows": matched_rows,
            "summaryRows": len(self.summary),
        }

    def expand_set(self) -> dict[str, Any]:
        if self.set_df is None:
            raise ValueError("ยังไม่ได้โหลดไฟล์ SET")
        if not self.summary:
            raise ValueError("ยังไม่มีผลคำนวณให้แตก SET")

        expanded: list[dict[str, Any]] = []
        set_found = 0
        set_missing: list[str] = []
        expanded_lines = 0

        set_lookup = self.set_df.copy()
        set_lookup["_set_key"] = set_lookup[SET_COL].astype(str).str.strip().str.lower()

        for row in self.summary:
            code = clean_text(row[CODE_COL])
            qty = parse_number(row[TOTAL_COL])
            key = code.lower()
            if key.startswith("set"):
                matches = set_lookup[set_lookup["_set_key"] == key]
                if matches.empty:
                    set_missing.append(code)
                    continue
                set_found += 1
                for _, item in matches.iterrows():
                    expanded.append(
                        {
                            MATERIAL_COL: clean_text(item[SET_DESC_COL]),
                            CODE_COL: clean_text(item[CODE_COL]),
                            TOTAL_COL: parse_number(item[SET_INSTALL_COL]) * qty,
                        }
                    )
                    expanded_lines += 1
            else:
                expanded.append(
                    {
                        MATERIAL_COL: clean_text(row[MATERIAL_COL]),
                        CODE_COL: code,
                        TOTAL_COL: qty,
                    }
                )

        self.summary = group_summary(expanded)
        return {
            "items": self.summary,
            "setFound": set_found,
            "setMissing": set_missing,
            "expandedLines": expanded_lines,
            "summaryRows": len(self.summary),
        }

    def export_summary(self) -> bytes:
        if not self.summary:
            raise ValueError("ยังไม่มีผลลัพธ์สำหรับ export")

        output = BytesIO()
        df = pd.DataFrame(self.summary, columns=[MATERIAL_COL, CODE_COL, TOTAL_COL])
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Summary")
        return output.getvalue()


def classify_wire_head(head: str) -> str | None:
    normalized = clean_text(head).upper()
    if normalized.startswith("DDE.BL"):
        return "dde_bl"
    if normalized.startswith("DDE"):
        return "dde"
    if normalized.startswith("DE"):
        return "de"
    if normalized.startswith("BA"):
        return "ba"
    return None


def validate_wire_selection(
    wire_kind: str | None,
    wire1: str,
    wire2: str,
    page_number: int,
    row_number: int,
    head: str,
) -> None:
    if wire_kind is None:
        return
    if wire1 not in WIRE_MATERIALS:
        raise ValueError(f"หน้า {page_number} แถว {row_number} ({head}): กรุณาเลือกชนิดสายช่องแรก")
    if wire_kind in {"dde", "dde_bl", "ba"} and wire2 not in WIRE_MATERIALS:
        raise ValueError(f"หน้า {page_number} แถว {row_number} ({head}): กรุณาเลือกชนิดสายช่องที่สอง")
    if wire_kind == "dde" and conductor_group(wire1) != conductor_group(wire2):
        raise ValueError(f"หน้า {page_number} แถว {row_number} ({head}): สายซ้ายและขวาต้องมีขนาดเดียวกันสำหรับ Tensionless")
    if wire_kind == "dde" and conductor_group(wire1) not in TENSIONLESS_MATERIALS:
        raise ValueError(f"หน้า {page_number} แถว {row_number} ({head}): ยังไม่มีรหัส Tensionless สำหรับสายขนาด {conductor_group(wire1)}")


def conductor_group(wire: str) -> str:
    details = WIRE_MATERIALS.get(clean_text(wire))
    return details[2] if details else ""


def add_material(
    totals: dict[tuple[str, str], dict[str, Any]],
    material: str,
    code: str,
    amount: float,
) -> None:
    if not code or amount == 0:
        return
    key = (material, code)
    if key not in totals:
        totals[key] = {MATERIAL_COL: material, CODE_COL: code, TOTAL_COL: 0.0}
    totals[key][TOTAL_COL] += amount


def add_wire_materials(
    totals: dict[tuple[str, str], dict[str, Any]],
    wire_kind: str | None,
    wire1: str,
    wire2: str,
    count: float,
) -> int:
    if wire_kind is None:
        return 0

    selected_wires = [wire1]
    if wire_kind in {"dde", "dde_bl"}:
        selected_wires.append(wire2)

    added = 0
    for wire in selected_wires:
        material, code, _, needs_clevis = WIRE_MATERIALS[wire]
        add_material(totals, material, code, 3 * count)
        added += 1
        if needs_clevis:
            clevis_material, clevis_code = CLEVIS_MATERIAL
            add_material(totals, clevis_material, clevis_code, 3 * count)
            added += 1

    if wire_kind == "dde":
        material, code = TENSIONLESS_MATERIALS[conductor_group(wire1)]
        tensionless_quantity = 3 * count
        add_material(totals, material, code, tensionless_quantity)
        added += 1
        for tape_material, tape_code in TENSIONLESS_TAPES:
            add_material(totals, tape_material, tape_code, tensionless_quantity)
            added += 1

    if wire_kind == "ba":
        if conductor_group(wire2) == "185":
            material, code = PG3_MATERIAL
            add_material(totals, material, code, 6 * count)
            added += 1
        else:
            for material, code in (HOTLINE_CLAMP_MATERIAL, BAIL_CLAMP_MATERIAL):
                add_material(totals, material, code, 3 * count)
                added += 1

    return added

    def export_page_summary(self, pages: list[list[dict[str, Any]]]) -> bytes:
        totals: dict[tuple[str, str], dict[int, float]] = {}
        for page_number, page in enumerate(pages, start=1):
            for item in page:
                size = clean_text(item.get("size"))
                head = clean_text(item.get("head"))
                count = parse_number(item.get("count"))
                if not size or not head or count <= 0:
                    continue
                key = (head, size)
                if key not in totals:
                    totals[key] = {}
                totals[key][page_number] = totals[key].get(page_number, 0.0) + count

        if not totals:
            raise ValueError("ยังไม่มีข้อมูลแต่ละหน้าสำหรับ export")

        page_columns = [f"หน้า {page_number}" for page_number in range(1, len(pages) + 1)]
        rows = []
        for (head, size), page_totals in sorted(totals.items(), key=lambda item: (natural_key(item[0][1]), natural_key(item[0][0]))):
            row: dict[str, Any] = {HEAD_COL: head, "เสา": size}
            for page_number, column in enumerate(page_columns, start=1):
                row[column] = page_totals.get(page_number)
            rows.append(row)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            columns = [HEAD_COL, "เสา", *page_columns]
            pd.DataFrame(rows, columns=columns).to_excel(writer, index=False, sheet_name="สรุปแต่ละหน้า", startrow=2)
            sheet = writer.sheets["สรุปแต่ละหน้า"]
            last_column = get_column_letter(len(columns))
            sheet.merge_cells(f"A1:{last_column}1")
            title = sheet["A1"]
            title.value = "สรุปรายการหัวเสาแยกตามหน้า"
            title.font = Font(name="Tahoma", size=16, bold=True, color="FFFFFF")
            title.fill = PatternFill("solid", fgColor="0F766E")
            title.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 28

            header_fill = PatternFill("solid", fgColor="DDEDEA")
            border = Border(bottom=Side(style="thin", color="AAB7C4"))
            for cell in sheet[3]:
                cell.font = Font(name="Tahoma", bold=True, color="1D242D")
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=len(columns)):
                for cell in row:
                    cell.font = Font(name="Tahoma", size=10)
                for cell in row[2:]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.###"

            sheet.column_dimensions["A"].width = 32
            sheet.column_dimensions["B"].width = 14
            for column_index in range(3, len(columns) + 1):
                sheet.column_dimensions[get_column_letter(column_index)].width = 13
            sheet.freeze_panes = "C4"
            sheet.auto_filter.ref = f"A3:{last_column}{sheet.max_row}"
            sheet.sheet_view.showGridLines = False

        return output.getvalue()


def read_set_sheet(path: str | Path) -> pd.DataFrame:
    sheets = pd.read_excel(path, sheet_name=None)
    if SET_SHEET in sheets:
        selected = sheets[SET_SHEET]
        if {SET_COL, CODE_COL, SET_DESC_COL, SET_INSTALL_COL}.issubset(set(selected.columns)):
            return selected
        normalized = normalize_report_set(selected)
        if normalized is not None:
            return normalized
    for df in sheets.values():
        if {SET_COL, CODE_COL, SET_DESC_COL, SET_INSTALL_COL}.issubset(set(df.columns)):
            return df
        normalized = normalize_report_set(df)
        if normalized is not None:
            return normalized
    first = next(iter(sheets.values()))
    return first


def normalize_report_set(df: pd.DataFrame) -> pd.DataFrame | None:
    source_set_col = "รหัสอุปกรณ์ต่อชุด"
    required = {source_set_col, CODE_COL, SET_DESC_COL, SET_INSTALL_COL}
    if not required.issubset(set(df.columns)):
        return None

    set_values = df[source_set_col].ffill()
    item_rows = df[source_set_col].isna() & df[CODE_COL].notna()
    return pd.DataFrame(
        {
            SET_COL: set_values[item_rows],
            CODE_COL: df.loc[item_rows, CODE_COL],
            SET_DESC_COL: df.loc[item_rows, SET_DESC_COL],
            SET_INSTALL_COL: df.loc[item_rows, SET_INSTALL_COL],
        }
    )


def require_columns(df: pd.DataFrame, columns: list[str], label: str) -> None:
    missing = [col for col in columns if col not in df.columns]
    if missing:
        raise ValueError(f"ไฟล์ {label} ขาดคอลัมน์: {', '.join(missing)}")


def group_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        material = clean_text(row[MATERIAL_COL])
        code = clean_text(row[CODE_COL])
        amount = parse_number(row[TOTAL_COL])
        if not code or amount == 0:
            continue
        key = (material, code)
        if key not in totals:
            totals[key] = {MATERIAL_COL: material, CODE_COL: code, TOTAL_COL: 0.0}
        totals[key][TOTAL_COL] += amount
    return sorted(totals.values(), key=lambda r: (str(r[CODE_COL]).lower(), str(r[MATERIAL_COL]).lower()))


def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def parse_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, str):
        expression = value.strip()
        if not expression or len(expression) > 100:
            return 0.0
        try:
            return float(evaluate_add_sub(ast.parse(expression, mode="eval").body))
        except (SyntaxError, TypeError, ValueError):
            return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def evaluate_add_sub(node: ast.AST) -> float:
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)) and not isinstance(node.value, bool):
        return float(node.value)
    if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub)):
        left = evaluate_add_sub(node.left)
        right = evaluate_add_sub(node.right)
        return left + right if isinstance(node.op, ast.Add) else left - right
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
        value = evaluate_add_sub(node.operand)
        return value if isinstance(node.op, ast.UAdd) else -value
    raise ValueError("รองรับเฉพาะตัวเลข เครื่องหมาย +, - และวงเล็บ")


def natural_key(value: str) -> list[tuple[int, Any]]:
    parts: list[tuple[int, Any]] = []
    current = ""
    numeric = False
    for char in str(value):
        is_digit = char.isdigit() or char == "."
        if current and is_digit != numeric:
            parts.append((0, float(current)) if numeric and current != "." else (1, current.lower()))
            current = char
        else:
            current += char
        numeric = is_digit
    if current:
        parts.append((0, float(current)) if numeric and current != "." else (1, current.lower()))
    return parts
